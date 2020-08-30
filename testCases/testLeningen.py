import time
from selenium import webdriver
from pageObjects.loginPage import Login
from pageObjects.LeningenPage import Leningen
from utilities.readProperties import configure
import openpyxl


class Test_002_Leningen:
    baseUrl = configure.getUrl()
    company = configure.getCompany()
    username = configure.getUsername()

    def test_NoRealtie(self, setup):
        self.driver = setup
        self.driver.get(self.baseUrl)
        self.driver.maximize_window()
        self.lp = Login(self.driver)
        self.lp.setCompany(self.company)
        # time.sleep(2)
        self.lp.setUsername(self.username)
        # time.sleep(1)
        self.lp.clickLogin()
        time.sleep(1)
        self.len = Leningen(self.driver)
        self.len.clickMenuFinanInstrumenten()
        time.sleep(3)
        self.len.clickMenuitemLeningen()
        time.sleep(2)
        self.len.clickNieuws()
        time.sleep(2)
        self.len.clickItemDrop("leningen")
        time.sleep(2)
        self.len.clickOk()
        time.sleep(5)
        self.len.clickLeningItem("Leningen")
        self.len.click_Geneer()
        self.len.set_NominaleWaarde(10000)
        self.len.click_RealtieExpand()
        time.sleep(3)
        number = self.len.set_RealatieNumber('10')
        # time.sleep(4)
        val = self.len.get_NoRelatie()
        if val == "Er zijn geen resultaten gevonden.":
            # self.len.click_1stCel()
            assert True
            self.driver.quit()
        else:
            self.driver.quit()
            assert False

    def test_CheckRelatie(self, setup):
        self.driver = setup
        self.driver.get(self.baseUrl)
        self.driver.maximize_window()
        self.lp = Login(self.driver)
        self.lp.setCompany(self.company)
        # time.sleep(2)
        self.lp.setUsername(self.username)
        # time.sleep(1)
        self.lp.clickLogin()
        time.sleep(1)
        self.len = Leningen(self.driver)
        self.len.clickMenuFinanInstrumenten()
        time.sleep(3)
        self.len.clickMenuitemLeningen()
        time.sleep(2)
        self.len.clickNieuws()
        time.sleep(2)
        self.len.clickItemDrop("leningen")
        time.sleep(2)
        self.len.clickOk()
        time.sleep(5)
        self.len.clickLeningItem("Leningen")
        self.len.click_Geneer()
        self.len.set_NominaleWaarde(10000)
        self.len.click_RealtieExpand()
        time.sleep(3)
        number = self.len.set_RealatieNumber('2')
        time.sleep(4)
        val = self.len.get_1stCel()
        if number == val:
            self.len.click_1stCel()
            assert True
            self.driver.quit()
        else:
            # print(mes)
            self.driver.quit()
            assert False

    def test_AllGoodFulling(self, setup):
        self.driver = setup
        self.driver.get(self.baseUrl)
        self.driver.maximize_window()
        self.lp = Login(self.driver)
        self.lp.setCompany(self.company)
        # time.sleep(2)
        self.lp.setUsername(self.username)
        # time.sleep(1)
        self.lp.clickLogin()
        time.sleep(1)
        self.len = Leningen(self.driver)
        self.len.clickMenuFinanInstrumenten()
        time.sleep(3)
        self.len.clickMenuitemLeningen()
        time.sleep(2)
        self.len.clickNieuws()
        time.sleep(2)
        self.len.clickItemDrop("leningen")
        time.sleep(2)
        self.len.clickOk()
        # Fulling
        time.sleep(5)
        self.len.clickLeningItem("Leningen")
        self.len.click_Geneer()
        self.len.set_NominaleWaarde(10000)
        self.len.click_RealtieExpand()
        time.sleep(3)
        number = self.len.set_RealatieNumber('2')
        time.sleep(4)
        self.len.click_1stCel()
        self.len.setStartDate('25-8-2020')
        self.len.clickSluiten()
        self.len.set_Looptijd(10)
        self.len.click_Opslaan()
        time.sleep(20)
        bol = self.len.get_WarningAutoIngevuld()
        if bol:
            assert bol, 'Click Opslaan'
            self.driver.quit()
        # assert bolean,"Goed test"
        # if str=="Eén of meerdere waarden zijn automatisch ingevuld. Controleer a.u.b. deze waarden en indien deze waarden correct zijn klik nogmaals op opslaan.":
        #    self.len.click_Opslaan()

    def test_WrongDateFormat(self, setup):
        self.driver = setup
        self.driver.get(self.baseUrl)
        self.driver.maximize_window()
        self.lp = Login(self.driver)
        self.lp.setCompany(self.company)
        # time.sleep(2)
        self.lp.setUsername(self.username)
        # time.sleep(1)
        self.lp.clickLogin()
        time.sleep(3)
        self.len = Leningen(self.driver)
        # self.len = Leningen(self.driver)
        self.len.clickMenuFinanInstrumenten()
        # time.sleep(3)
        self.len.clickMenuitemLeningen()
        # time.sleep(2)
        self.len.clickNieuws()
        self.len.clickItemDrop("leningen")

        time.sleep(2)
        self.len.clickOk()
        path = "D:\Python Projects\TreasuryProject\TestData\LenWrongDateFormat.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet1 = workbook.get_sheet_by_name("Blad1")
        self.len.clickLeningItem(sheet1.cell(2, 1).value)
        if sheet1.cell(2, 2).value == None:
            self.len.click_Geneer()
        self.len.set_NominaleWaarde(sheet1.cell(2, 3).value)
        self.len.click_RealtieExpand()
        time.sleep(3)
        number = self.len.set_RealatieNumber(sheet1.cell(2, 3).value)
        time.sleep(3)
        self.len.click_1stCel()
        self.len.setStartDate(sheet1.cell(2, 4).value)
        self.len.clickSluiten()
        self.len.set_Looptijd(10)
        self.len.click_Opslaan()
        time.sleep(20)
        bol = self.len.get_WarningDateForm()
        if bol:
            assert bol, 'Niet goed ingangsdatum'
            self.driver.quit()
